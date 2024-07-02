import openpyxl
import os
import glob

for index in range(1000, 1013):  # 这个循环只会执行一次
    try:
        file_pattern = os.path.join('original_data/Shanghai_T1DM/', f'{index}*.xlsx')
        all_files = glob.glob(file_pattern)

        # 提取文件名
        all_file_names = [os.path.basename(file) for file in all_files]

        # 打印文件名
        for filename in all_file_names:
            try:
                print(filename)
                workbook = openpyxl.load_workbook("original_data/Shanghai_T1DM/" + filename)

                # 获取第一个工作表（这里只有一个）
                sheet = workbook.active

                # 获取合并单元格
                merged_cells = sheet.merged_cells

                # 强制加入11列，防止出现空表（无合并单元格）情况
                sheet.cell(1, 12).value = "CSII - basal insulin (Novolin R, IU / H)"

                # 遍历每一个合并单元格
                for merged_cell in merged_cells:
                    start_row, start_column, end_row, end_column = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col

                    # 获取合并单元格的值
                    merged_value = sheet.cell(start_row, start_column).value

                    # 拆分合并单元格，并填充值
                    for row in range(start_row, end_row + 1):
                        for col in range(start_column, end_column + 1):
                            # 建立一个新列，把数据放进去
                            sheet.cell(row, col + 2).value = merged_value

                # 处理合并单元格间的单个单元格
                for row in sheet.iter_rows():
                    cell = row[11]  # 新列11
                    if cell.value is None:
                        cell.value = row[9].value
                    # 暂停时全部置为0
                    if cell.value == "temporarily suspend insulin delivery":
                        cell.value = 0

                workbook.save("generated_data/Shanghai_T1DM/" + filename)
            except Exception as e:
                print(f"Error processing file {filename}: {e}")
    except Exception as e:
        print(f"Error during processing index {index}: {e}")
